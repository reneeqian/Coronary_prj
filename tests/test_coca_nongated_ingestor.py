"""
Tests for COCANongatedIngestor.

All synthetic tests build minimal in-memory DICOM and xlsx fixtures so no real
dataset is required.  The FakeDicom helper mirrors the same pattern used in
test_coca_ingestor_synthetic.py for the gated ingestor.
"""

import math
from unittest.mock import patch

import numpy as np
import openpyxl
import pytest
from regulatory_tools.evidence.evidence_report import EvidenceReport

from Coronary_prj.ingestors.coca_gated_ingestor import DatasetStructureError
from Coronary_prj.ingestors.coca_nongated_ingestor import COCANongatedIngestor

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


class FakeDicom:
    def __init__(self, z: float, pixel_value: float = 0.0):
        self.ImagePositionPatient = [0.0, 0.0, float(z)]
        self.PixelSpacing = [0.5, 0.5]
        self.SliceThickness = 1.0
        self.RescaleSlope = 1.0
        self.RescaleIntercept = 0.0
        self.pixel_array = np.full((4, 4), pixel_value, dtype=np.float32)


def _write_scores_xlsx(path, rows: list[tuple]) -> None:
    """Write a minimal scores.xlsx with header + data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["filename ", "LCA", "LAD", "LCX", "RCA", "total"])
    for row in rows:
        ws.append(list(row))
    wb.save(path)


def _make_patient_dir(root, patient_id: str, n_dcm: int = 1) -> None:
    """Create {root}/{patient_id}/ with n_dcm placeholder .dcm files."""
    d = root / patient_id
    d.mkdir(parents=True, exist_ok=True)
    for i in range(n_dcm):
        (d / f"IM-{i:04d}.dcm").write_text("fake")


# ---------------------------------------------------------------------------
# Dataset structure validation (DAT-015)
# ---------------------------------------------------------------------------


@pytest.mark.requirement("DAT-015")
def test_scores_file_not_found_raises(tmp_path, evidence_output_dir):
    """Ingestor constructor raises DatasetStructureError when scores.xlsx is absent."""
    report = EvidenceReport(subject="DAT-015: missing scores.xlsx raises DatasetStructureError")
    _make_patient_dir(tmp_path, "1")
    raised = False
    try:
        COCANongatedIngestor(tmp_path)
    except DatasetStructureError:
        raised = True
    if not raised:
        report.error(
            "Expected DatasetStructureError when scores.xlsx absent; none raised", "DAT-015"
        )
    report.info(
        "DatasetStructureError raised on construction when scores.xlsx is absent", "DAT-015"
    )
    report.auto_save("DAT015_scores_file_not_found", evidence_output_dir)
    assert not report.has_errors, report.summary()


@pytest.mark.requirement("DAT-015")
def test_empty_dataset_root_raises(tmp_path, evidence_output_dir):
    """list_patient_ids raises DatasetStructureError when root has no subdirectories."""
    report = EvidenceReport(
        subject="DAT-015: list_patient_ids raises when root has no patient directories"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])

    ingestor = COCANongatedIngestor(tmp_path)
    with pytest.raises(DatasetStructureError):
        ingestor.list_patient_ids()

    report.info("DatasetStructureError raised when dataset root has no subdirectories", "DAT-015")
    report.auto_save("dat015_empty_dataset_root_raises", evidence_output_dir)
    assert not report.has_errors, report.summary()


@pytest.mark.requirement("DAT-015")
def test_missing_patient_directory_raises(tmp_path, evidence_output_dir):
    """load_patient_sample raises DatasetStructureError for a non-existent patient dir."""
    report = EvidenceReport(
        subject="DAT-015: load_patient_sample raises for non-existent patient directory"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("99A", 0, 0, 0, 0, 0)])
    _make_patient_dir(tmp_path, "1")

    ingestor = COCANongatedIngestor(tmp_path)
    with pytest.raises(DatasetStructureError):
        ingestor.load_patient_sample("99")

    report.info(
        "DatasetStructureError raised for patient id '99' with no matching directory", "DAT-015"
    )
    report.auto_save("dat015_missing_patient_directory_raises", evidence_output_dir)
    assert not report.has_errors, report.summary()


@pytest.mark.requirement("DAT-015")
def test_no_dicom_files_in_patient_dir_raises(tmp_path, evidence_output_dir):
    """load_patient_sample raises DatasetStructureError if patient dir has no .dcm files."""
    report = EvidenceReport(
        subject="DAT-015: load_patient_sample raises when patient dir has no .dcm files"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    (tmp_path / "1").mkdir()

    ingestor = COCANongatedIngestor(tmp_path)
    with pytest.raises(DatasetStructureError):
        ingestor.load_patient_sample("1")

    report.info(
        "DatasetStructureError raised when patient directory contains no .dcm files", "DAT-015"
    )
    report.auto_save("dat015_no_dicom_files_raises", evidence_output_dir)
    assert not report.has_errors, report.summary()


@pytest.mark.requirement("DAT-015")
def test_scores_xlsx_missing_required_column_raises(tmp_path):
    """Constructor raises DatasetStructureError when scores.xlsx has wrong columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["filename", "LCA", "LAD"])  # missing LCX, RCA, total
    wb.save(tmp_path / "scores.xlsx")

    with pytest.raises(DatasetStructureError, match="missing column"):
        COCANongatedIngestor(tmp_path)


# ---------------------------------------------------------------------------
# Score loading (DAT-016)
# ---------------------------------------------------------------------------


@pytest.mark.requirement("DAT-016")
def test_scores_attached_to_patient_metadata(tmp_path, evidence_output_dir):
    """Patient metadata contains correct per-vessel scores from scores.xlsx."""
    report = EvidenceReport(subject="COCANongatedIngestor score loading")

    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 132.0, 1.91, 128.0, 37.5, 300.0)])
    _make_patient_dir(tmp_path, "1")

    with patch("pydicom.dcmread", return_value=FakeDicom(0)):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("1")

    meta = sample.metadata
    if abs(meta.get("lca", -1) - 132.0) > 1e-6:
        report.error(f"lca wrong: got {meta.get('lca')}", "DAT-016")
    if abs(meta.get("lad", -1) - 1.91) > 1e-4:
        report.error(f"lad wrong: got {meta.get('lad')}", "DAT-016")
    if abs(meta.get("lcx", -1) - 128.0) > 1e-6:
        report.error(f"lcx wrong: got {meta.get('lcx')}", "DAT-016")
    if abs(meta.get("rca", -1) - 37.5) > 1e-6:
        report.error(f"rca wrong: got {meta.get('rca')}", "DAT-016")
    if abs(meta.get("total_score", -1) - 300.0) > 1e-6:
        report.error(f"total_score wrong: got {meta.get('total_score')}", "DAT-016")

    report.info(
        f"Per-vessel scores loaded from scores.xlsx: lca={meta['lca']}, lad={meta['lad']}, "
        f"lcx={meta['lcx']}, rca={meta['rca']}, total={meta['total_score']}",
        "DAT-016",
    )
    report.auto_save("DAT016_scores_attached", evidence_output_dir)
    assert not report.has_errors, report.summary()


@pytest.mark.requirement("DAT-016")
def test_missing_score_entry_uses_zero_fill(tmp_path, evidence_output_dir):
    """Patient with no score row in xlsx gets zero-filled scores without raising."""
    report = EvidenceReport(
        subject="DAT-016: patient absent from scores.xlsx gets zero-filled scores"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    _make_patient_dir(tmp_path, "103")

    with patch("pydicom.dcmread", return_value=FakeDicom(0)):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("103")

    report.info(
        f"Patient 103 (absent from xlsx): total_score={sample.metadata['total_score']}, lca={sample.metadata['lca']}",
        "DAT-016",
    )
    report.auto_save("dat016_missing_score_entry_zero_fill", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert sample.metadata["total_score"] == 0.0
    assert sample.metadata["lca"] == 0.0


@pytest.mark.requirement("DAT-016")
def test_missing_score_entry_warns_via_report(tmp_path, evidence_output_dir):
    """Missing score entry is reported via the EvidenceReport warn channel."""
    ev_report = EvidenceReport(
        subject="DAT-016: missing score entry produces a warning on the ingestor report"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    _make_patient_dir(tmp_path, "2")

    class CapturingReport:
        def __init__(self):
            self.warnings = []

        def warn(self, message, requirement_tag=None, context=None):
            self.warnings.append(message)

        def info(self, message, requirement_tag=None, context=None):
            pass

    capturing = CapturingReport()
    with patch("pydicom.dcmread", return_value=FakeDicom(0)):
        ingestor = COCANongatedIngestor(tmp_path, report=capturing)
        ingestor.load_patient_sample("2")

    has_warning = any("2" in w for w in capturing.warnings)
    if not has_warning:
        ev_report.error("No warning emitted for patient '2' absent from scores.xlsx", "DAT-016")
    else:
        ev_report.info(
            f"Warning emitted for patient '2' missing from scores.xlsx: {capturing.warnings}",
            "DAT-016",
        )

    ev_report.auto_save("dat016_missing_score_warns", evidence_output_dir)
    assert not ev_report.has_errors, ev_report.summary()
    assert has_warning, "Expected warning for missing score"


@pytest.mark.requirement("DAT-016")
def test_blank_score_cell_treated_as_zero(tmp_path, evidence_output_dir):
    """None/blank cells in scores.xlsx are coerced to 0.0."""
    report = EvidenceReport(subject="DAT-016: blank/None score cells in xlsx are coerced to 0.0")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["filename ", "LCA", "LAD", "LCX", "RCA", "total"])
    ws.append(["5A", None, None, 0, None, None])
    wb.save(tmp_path / "scores.xlsx")

    _make_patient_dir(tmp_path, "5")

    with patch("pydicom.dcmread", return_value=FakeDicom(0)):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("5")

    report.info(
        f"Blank cells coerced: lca={sample.metadata['lca']}, total_score={sample.metadata['total_score']}",
        "DAT-016",
    )
    report.auto_save("dat016_blank_score_cell_treated_as_zero", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert sample.metadata["lca"] == 0.0
    assert sample.metadata["total_score"] == 0.0


# ---------------------------------------------------------------------------
# DICOM loading / ingestion (DAT-004, DAT-012)
# ---------------------------------------------------------------------------


@pytest.mark.requirement("DAT-004")
def test_slices_z_sorted_on_load(tmp_path, evidence_output_dir):
    """Slices are sorted by ImagePositionPatient[2] before stacking."""
    report = EvidenceReport(subject="COCANongatedIngestor Z-sort")

    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    patient_dir = tmp_path / "1"
    patient_dir.mkdir()

    files = [patient_dir / f"IM-{i}.dcm" for i in range(3)]
    for f in files:
        f.write_text("fake")

    fake_dicoms = {
        str(files[0]): FakeDicom(10, 10),
        str(files[1]): FakeDicom(5, 5),
        str(files[2]): FakeDicom(7, 7),
    }

    def fake_dcmread(path, *args, **kwargs):
        return fake_dicoms[str(path)]

    with patch("pydicom.dcmread", side_effect=fake_dcmread):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("1")

    vol = sample.image_volume
    if vol[0, 0, 0] != 5 or vol[1, 0, 0] != 7 or vol[2, 0, 0] != 10:
        report.error(f"Unexpected Z order: {vol[:, 0, 0]}", "DAT-004")

    report.info(
        f"Nongated slices sorted by Z: [{vol[0, 0, 0]:.0f}, {vol[1, 0, 0]:.0f}, {vol[2, 0, 0]:.0f}] = [5,7,10]",
        "DAT-004",
    )
    report.auto_save("DAT004_nongated_z_sort", evidence_output_dir)
    assert not report.has_errors, report.summary()


@pytest.mark.requirement("DAT-004")
def test_hounsfield_rescale_applied(tmp_path, evidence_output_dir):
    """RescaleSlope and RescaleIntercept are applied to pixel values."""
    report = EvidenceReport(
        subject="DAT-004: RescaleSlope and RescaleIntercept are applied to nongated CT pixel values"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    _make_patient_dir(tmp_path, "1")

    class RescaleDicom:
        ImagePositionPatient = [0, 0, 0]
        PixelSpacing = [1.0, 1.0]
        SliceThickness = 1.0
        RescaleSlope = 2.0
        RescaleIntercept = -1000.0
        pixel_array = np.ones((2, 2), dtype=np.float32) * 100

    with patch("pydicom.dcmread", return_value=RescaleDicom()):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("1")

    expected = -800.0
    actual = float(sample.image_volume[0, 0, 0])
    report.info(
        f"HU rescale: pixel=100 * slope=2 + intercept=-1000 → {actual} (expected {expected})",
        "DAT-004",
    )
    report.auto_save("dat004_nongated_hounsfield_rescale_applied", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert np.all(sample.image_volume == expected)


@pytest.mark.requirement("DAT-004")
def test_annotations_always_none(tmp_path, evidence_output_dir):
    """Nongated ingestor never produces spatial annotations."""
    report = EvidenceReport(
        subject="DAT-004: nongated ingestor always returns None for vector_rois (score-only dataset)"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 100, 200, 50, 75, 425)])
    _make_patient_dir(tmp_path, "1")

    with patch("pydicom.dcmread", return_value=FakeDicom(0)):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("1")

    report.info(
        f"annotations.vector_rois={sample.annotations.vector_rois!r} (expected None)", "DAT-004"
    )
    report.auto_save("dat004_nongated_annotations_always_none", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert sample.annotations.vector_rois is None


@pytest.mark.requirement("DAT-015")
def test_missing_image_position_patient_warns_and_skips(tmp_path, evidence_output_dir):
    """DICOMs without ImagePositionPatient are skipped with a warning."""
    report = EvidenceReport(
        subject="DAT-015: DICOMs missing ImagePositionPatient are skipped; remaining slices still loaded"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    patient_dir = tmp_path / "1"
    patient_dir.mkdir()

    files = [patient_dir / f"IM-{i}.dcm" for i in range(2)]
    for f in files:
        f.write_text("fake")

    class DicomNoPosition:
        PixelSpacing = [1.0, 1.0]
        SliceThickness = 1.0
        RescaleSlope = 1.0
        RescaleIntercept = 0.0
        pixel_array = np.zeros((2, 2), dtype=np.float32)

    class DicomWithPosition:
        ImagePositionPatient = [0, 0, 0]
        PixelSpacing = [1.0, 1.0]
        SliceThickness = 1.0
        RescaleSlope = 1.0
        RescaleIntercept = 0.0
        pixel_array = np.zeros((2, 2), dtype=np.float32)

    call_map = {str(files[0]): DicomNoPosition(), str(files[1]): DicomWithPosition()}

    def fake_dcmread(path, *args, **kwargs):
        return call_map[str(path)]

    with patch("pydicom.dcmread", side_effect=fake_dcmread):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("1")

    report.info(
        f"1 DICOM missing ImagePositionPatient skipped; volume.shape[0]={sample.image_volume.shape[0]} (expected 1)",
        "DAT-015",
    )
    report.auto_save("dat015_missing_image_position_skipped", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert sample.image_volume.shape[0] == 1


@pytest.mark.requirement("DAT-015")
def test_spacing_fallback_on_missing_dicom_metadata(tmp_path, evidence_output_dir):
    """Missing PixelSpacing falls back to (1,1,1) without raising."""
    report = EvidenceReport(
        subject="DAT-015: missing PixelSpacing metadata falls back to (1,1,1) without raising"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    _make_patient_dir(tmp_path, "1")

    class DicomNoSpacing:
        ImagePositionPatient = [0, 0, 0]
        RescaleSlope = 1.0
        RescaleIntercept = 0.0
        pixel_array = np.zeros((2, 2), dtype=np.float32)

    with patch("pydicom.dcmread", return_value=DicomNoSpacing()):
        ingestor = COCANongatedIngestor(tmp_path)
        sample = ingestor.load_patient_sample("1")

    report.info(f"Spacing fallback: spacing={sample.spacing} (expected (1.0, 1.0, 1.0))", "DAT-015")
    report.auto_save("dat015_spacing_fallback_on_missing_metadata", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert sample.spacing == (1.0, 1.0, 1.0)


# ---------------------------------------------------------------------------
# list_patient_ids ordering (DAT-002)
# ---------------------------------------------------------------------------


@pytest.mark.requirement("DAT-002")
def test_list_patient_ids_sorted_numerically(tmp_path, evidence_output_dir):
    """list_patient_ids returns patient dirs sorted numerically."""
    report = EvidenceReport(
        subject="DAT-002: list_patient_ids returns patient ids in numeric sort order"
    )
    _write_scores_xlsx(
        tmp_path / "scores.xlsx",
        [
            ("1A", 0, 0, 0, 0, 0),
            ("10A", 0, 0, 0, 0, 0),
            ("2A", 0, 0, 0, 0, 0),
        ],
    )
    for pid in ["1", "2", "10"]:
        (tmp_path / pid).mkdir()

    ingestor = COCANongatedIngestor(tmp_path)
    ids = ingestor.list_patient_ids()

    report.info(f"list_patient_ids() returned {ids!r} (expected ['1', '2', '10'])", "DAT-002")
    report.auto_save("dat002_list_patient_ids_sorted_numerically", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert ids == ["1", "2", "10"]


# ---------------------------------------------------------------------------
# get_sample API (DAT-016)
# ---------------------------------------------------------------------------


@pytest.mark.requirement("DAT-016")
def test_get_sample_returns_volume_and_score_array(tmp_path, evidence_output_dir):
    """get_sample returns (volume_array, scores_array) with shape (4,)."""
    report = EvidenceReport(
        subject="DAT-016: get_sample returns numpy volume and score array with shape (4,)"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 10.0, 20.0, 30.0, 40.0, 100.0)])
    _make_patient_dir(tmp_path, "1")

    with patch("pydicom.dcmread", return_value=FakeDicom(0)):
        ingestor = COCANongatedIngestor(tmp_path)
        volume, scores = ingestor.get_sample("1")

    report.info(
        f"get_sample('1'): volume type={type(volume).__name__}, scores.shape={scores.shape}, "
        f"lca={scores[0]:.2f}, lad={scores[1]:.2f}",
        "DAT-016",
    )
    report.auto_save("dat016_get_sample_volume_and_score_array", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert isinstance(volume, np.ndarray)
    assert scores.shape == (4,)
    assert abs(scores[0] - 10.0) < 1e-5  # lca
    assert abs(scores[1] - 20.0) < 1e-5  # lad


# ---------------------------------------------------------------------------
# Determinism (DAT-008)
# ---------------------------------------------------------------------------


@pytest.mark.requirement("DAT-008")
def test_deterministic_patient_loading(tmp_path, evidence_output_dir):
    """Loading the same patient twice returns identical volumes."""
    report = EvidenceReport(
        subject="DAT-008: loading the same nongated patient twice returns identical volumes"
    )
    _write_scores_xlsx(tmp_path / "scores.xlsx", [("1A", 0, 0, 0, 0, 0)])
    _make_patient_dir(tmp_path, "1")

    with patch("pydicom.dcmread", return_value=FakeDicom(0, 42.0)):
        ingestor = COCANongatedIngestor(tmp_path)
        s1 = ingestor.load_patient_sample("1")
        s2 = ingestor.load_patient_sample("1")

    if not np.array_equal(s1.image_volume, s2.image_volume):
        report.error("Two consecutive loads returned different image volumes", "DAT-008")
    if s1.metadata["total_score"] != s2.metadata["total_score"]:
        report.error("Two consecutive loads returned different total_score", "DAT-008")
    report.info(
        "Two consecutive load_patient_sample() calls returned identical volume and metadata",
        "DAT-008",
    )
    report.auto_save("DAT008_nongated_deterministic_loading", evidence_output_dir)
    assert not report.has_errors, report.summary()
    assert np.array_equal(s1.image_volume, s2.image_volume)
    assert s1.metadata["total_score"] == s2.metadata["total_score"]
